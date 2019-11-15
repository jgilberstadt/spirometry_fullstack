# subclass for generating the monthly report

# import libraries for PostgreSQL connection
from dbconfig import config
import psycopg2
# import libraries for loading workbook templates
from openpyxl import load_workbook
# import the basic class
from basic_report import BasicReport

import matplotlib.pyplot as plt
import numpy as np

class MonthlyReport(BasicReport):
	def __init__(self):
		super().__init__(db_config_path='../credential/database.ini', 
			workbook_path='Report_Templates/Report template - R6.xlsx', report_type=0)

		# put report-specific initalization here
		# e.g. graph objects, replaced values queried from DB

	# Plot the graphs and place them in the spreedsheet
	def plotGraph(self):
		weeks = self.test_date_list
		fevs = self.fev1max_list
		minWeek = np.amin(weeks)
		maxWeek = np.amax(weeks)
		minFev = np.amin(fevs)
		maxFev = np.amax(fevs)

		cursor = self.cnx.cursor()
		cursor.execute("SELECT nl_upper, nl_lower FROM patient_data WHERE imei_num=%s", ([self.imei_num]))
		for (nl_upper, nl_lower) in cursor:
			nLower = nl_lower
			nUpper = nl_upper
		

		percent = round((1.0 - (minFev/maxFev)), 2)
		
		plt.axhline(y=nLower, color='r', linestyle='-')
		plt.axhline(y=nUpper, color='r', linestyle='-', label='Upper/Lower NL')

		plt.xticks(np.arange(minWeek, maxWeek, 5))
		plt.yticks(np.arange(minFev, maxFev, 100))

		plt.xlabel('Weeks')
		plt.ylabel('FEV1 (mL)')
		plt.suptitle('Monitoring Weeks (Last 20 Weeks)')
		plt.grid(True)

		coef = np.polyfit(weeks,fevs,1)
		poly1d_fn = np.poly1d(coef) 
		# poly1d_fn is now a function which takes in x and returns an estimate for y

		plt.plot(weeks,fevs, 'ro', label='FEV')
		plt.plot(weeks, poly1d_fn(weeks), '-k', label='Regression')

		plt2 = plt.twinx()
		plt2.set_ylim(percent, 0)
		plt2.set_ylabel('Percent of Max')

		# self.ws['E3'] = report_date.strftime("%m/%d/%y")

		plt.legend(loc='best')
		plotName = self.imei_num+"_monthly_Report_"+str(self.current_month)+".png"
		plt.savefig(plotName)
		img = openpyxl.drawing.Image(plotName)
		img.anchor(self.ws.cell('P1'))

		self.ws.add_image(img)
		# plt.show()
		


	# Query the PostgreSQL for personalized values to replace the values in the report template
	def replaceValues(self):

# If everything is taken care by the above two methods, we can just save the xlsx in this method
	def generateReport(self):