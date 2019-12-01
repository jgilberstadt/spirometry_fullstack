# basic class for generating report

# import libraries for PostgreSQL connection
from dbconfig import config
import psycopg2
# import libraries for loading workbook templates
from openpyxl import load_workbook

class BasicReport:
	def __init__(self, db_config_path=None, workbook_path=None, report_type=0):
		# initialize the database connection 
		# connect to PostgreSQL database
		self.db_config_path = db_config_path
		# get the database credentials from the .ini config file
		dbc = config('database.ini' if self.db_config_path == None else self.db_config_path, 'postgresql')
		self.cnx = psycopg2.connect(**dbc)

		# initialize the workbook template that can be used for populating the values obtained 
		# from the PostgreSQL table
		self.workbook_path = 'Report_Templates/Report template - R6.xlsx' if workbook_path == None else workbook_path
		self.wb = load_workbook(self.workbook_path)

		# initialize the report type which can be used for selecting the corresponding worksheet.
		# 0: Report Datasheet (Monthly)
		# 1: Report Datasheet M (V+AMI-P+)
		# 2: Report Datasheet (V+AMI-P+)
		# 3: Report Datasheet (V+AMI-P-)
		# 4: Report Datasheet (V+AMI+-)
		# 5: Report Datasheet (Conv 10 wks)
		# 6: Report Datasheet (Conv 6 wks)
		self.report_type_dict = {0:'Report Datasheet (Monthly)', 1:'Report Datasheet M (V+AMI-P+)',
		2:'Report Datasheet (V+AMI-P+)', 3:'Report Datasheet (V+AMI-P-)', 4:'Report Datasheet (V+AMI+-)',
		5:'Report Datasheet (Conv 10 wks)', 6:'Report Datasheet (Conv 6 wks)'}
		self.report_type_name = self.report_type_dict[report_type]

	def empty_row(self, index):
		self.ws['A'+str(index)] = ""
		self.ws['B'+str(index)] = ""
		self.ws['C'+str(index)] = ""
		self.ws['D'+str(index)] = ""
		self.ws['E'+str(index)] = ""
		self.ws['F'+str(index)] = ""
		self.ws['G'+str(index)] = ""
		self.ws['N'+str(index)] = ""
		self.ws['O'+str(index)] = ""
		self.ws['S'+str(index)] = ""

	def empty_row_all(self, index):
		self.ws['A'+str(index)] = ""
		self.ws['B'+str(index)] = ""
		self.ws['C'+str(index)] = ""
		self.ws['D'+str(index)] = ""
		self.ws['E'+str(index)] = ""
		self.ws['F'+str(index)] = ""
		self.ws['G'+str(index)] = ""
		self.ws['H'+str(index)] = ""
		self.ws['I'+str(index)] = ""
		self.ws['J'+str(index)] = ""
		self.ws['K'+str(index)] = ""
		self.ws['L'+str(index)] = ""
		self.ws['M'+str(index)] = ""
		self.ws['N'+str(index)] = ""
		self.ws['O'+str(index)] = ""
		self.ws['P'+str(index)] = ""
		self.ws['Q'+str(index)] = ""
		self.ws['R'+str(index)] = ""
		self.ws['S'+str(index)] = ""

	# Plot the graphs and place them in the spreedsheet
	def plotGraph(self):
		print("This is the basic class method, please implement it in the subclass")

	# Query the PostgreSQL for personalized values to replace the values in the report template
	def replaceValues(self):
		print("This is the basic class method, please implement it in the subclass")

	# If everything is taken care by the above two methods, we can just save the xlsx in this method
	def generateReport(self):
		print("This is the basic class method, please implement it in the subclass")



