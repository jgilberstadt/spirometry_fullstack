# subclass for generating the monthly report

# import libraries for PostgreSQL connection
from dbconfig import config
# import libraries for loading workbook templates
from openpyxl import load_workbook
# import the basic class
from basic_report import BasicReport

import matplotlib.pyplot as plt
import numpy as np
import psycopg2
from datetime import datetime, timedelta, date

class SurveillanceReport(BasicReport):
	def __init__(self, imei_num):
		BasicReport.__init__(self, db_config_path='credentials/database.ini', 
			workbook_path='Report_Templates/Report template - R6.xlsx', report_type=0)

		# put report-specific initalization here
		# e.g. graph objects, replaced values queried from DB
		self.imei_num = imei_num
		#self.current_month = current_month
		self.ws = self.wb['Report Datasheet (V+AMI-P+)']

	# Plot the graphs and place them in the spreedsheet
	def plotGraph(self):
		weeks = self.test_date_list
		fevs = self.fev1max_list
		minWeek = np.amin(weeks)
		maxWeek = np.amax(weeks)
		minFev = np.amin(fevs)
		maxFev = np.amax(fevs)

		cursor = self.cnx.cursor()
		cursor.execute("SELECT nl_upper, nl_lower FROM patient_data WHERE imei_num=%s", ([self.imei_num]))
		for (nl_upper, nl_lower) in cursor:
			nLower = nl_lower
			nUpper = nl_upper
		

		percent = round((1.0 - (minFev/maxFev)), 2)
		
		plt.axhline(y=nLower, color='r', linestyle='-')
		plt.axhline(y=nUpper, color='r', linestyle='-', label='Upper/Lower NL')

		plt.xticks(np.arange(minWeek, maxWeek, 5))
		plt.yticks(np.arange(minFev, maxFev, 100))

		plt.xlabel('Weeks')
		plt.ylabel('FEV1 (mL)')
		plt.suptitle('Monitoring Weeks (Last 20 Weeks)')
		plt.grid(True)

		coef = np.polyfit(weeks,fevs,1)
		poly1d_fn = np.poly1d(coef) 
		# poly1d_fn is now a function which takes in x and returns an estimate for y

		plt.plot(weeks,fevs, 'ro', label='FEV')
		plt.plot(weeks, poly1d_fn(weeks), '-k', label='Regression')

		plt2 = plt.twinx()
		plt2.set_ylim(percent, 0)
		plt2.set_ylabel('Percent of Max')

		# self.ws['E3'] = report_date.strftime("%m/%d/%y")

		plt.legend(loc='best')
		plotName = self.imei_num+"_Surveillance_Report_"+str(self.current_month)+".png"
		plt.savefig(plotName)
		img = openpyxl.drawing.Image(plotName)
		img.anchor(self.ws.cell('P1'))

		self.ws.add_image(img)
		# plt.show()

	# Query the PostgreSQL for personalized values to replace the values in the report template
	def replaceValues(self):
		cursor = self.cnx.cursor()
		# obtain current datetime
		#report_date = datetime.now()
		report_date = date(2019,11,4)
		self.ws['E3'] = report_date.strftime("%m/%d/%y")
		# query patient study number and other information that can be used later
		cursor.execute("SELECT imei_num, start_study_date, date_of_transplant, mode FROM patient_data WHERE imei_num=%s", ([self.imei_num]))
		for (patient_study_number, start_study_date, date_of_transplant, mode) in cursor:
			self.ws['E2'] = patient_study_number
			break

		# create a list of fev1max values to be used in replaceValues(self) function
		self.fev1max_list = []
		# create a list of date to be used in replaceValues(self) function
		self.test_date_list = []

		# the raw data extraction for the conversion report is a bit tricky
		# first, need to empty the raw data in the template
		# then, align dates with rows

		# the alignment cursor
		# at the begining, set the alignment cursor to be 72 days before, which is the start date of the study
		alignment_cursor = report_date - timedelta(days=72)

		# query all fev1 values from spiro_data table
		raw_data_incremental_index = 45

		cursor.execute("SELECT fev11, fev12, fev13, fev14, fev15, fev16, test_date, is_variance, variance_test_counter FROM spiro_data WHERE imei_num=%s AND test_date >=%s AND test_date <%s", ([self.imei_num, alignment_cursor, report_date]))
		for (fev11, fev12, fev13, fev14, fev15, fev16, test_date, is_variance, variance_test_counter) in cursor:
			self.empty_row(raw_data_incremental_index)
			self.ws['A'+str(raw_data_incremental_index)] = int(patient_study_number)
			self.ws['B'+str(raw_data_incremental_index)] = fev11
			self.ws['C'+str(raw_data_incremental_index)] = fev12
			self.ws['D'+str(raw_data_incremental_index)] = fev13
			self.ws['E'+str(raw_data_incremental_index)] = fev14
			self.ws['F'+str(raw_data_incremental_index)] = fev15
			self.ws['G'+str(raw_data_incremental_index)] = fev16
			self.ws['N'+str(raw_data_incremental_index)] = test_date
			self.ws['O'+str(raw_data_incremental_index)] = start_study_date
			self.ws['Q'+str(raw_data_incremental_index)] = date_of_transplant
			self.fev1max_list.append(max(float(fev11), float(fev12), float(fev13), float(fev14), float(fev15), float(fev16)))
			self.test_date_list.append(test_date)


			# need to associate each session with mode
			if mode == 1:
			   self.ws['S'+str(raw_data_incremental_index)] = 'Pre-surv'
			elif mode == 2 or mode == 3:
			   self.ws['S'+str(raw_data_incremental_index)] = 'Month'
			#a
			raw_data_incremental_index += 1
			#alignment_cursor += timedelta(days=1)

		print "raw_data_incremental_index", raw_data_incremental_index

		while(raw_data_incremental_index<=116):
			self.empty_row_all(raw_data_incremental_index)
			raw_data_incremental_index += 1

		cursor.execute("SELECT maxrate, minrate, lowestsat FROM pulse_data WHERE imei_num=%s AND test_date BETWEEN current_date - 30 and current_date", ([self.imei_num]))
		currentminrate = 1000
		currentlowestsat = 1000
		currentmaxrate = 0
		for (maxrate, minrate, lowestsat) in cursor:
			if minrate < currentminrate and minrate > 30:
				currentminrate = minrate
			if lowestsat < currentlowestsat:
				currentlowestsat = lowestsat
			if maxrate > currentmaxrate and maxrate > 30:
				currentmaxrate = maxrate

		"""
		print(currentminrate)
		print(currentmaxrate)
		print(lowestsat)
		"""
		self.ws['I11'] = str(currentminrate)+'-'+str(currentmaxrate)
		self.ws['I8'] = lowestsat

		#O2 Mean Sat
		cursor.execute("SELECT lowestsat FROM pulse_data WHERE imei_num=%s AND test_date BETWEEN current_date - 30 and current_date", ([self.imei_num]))
		totalSaturation = 0
		numberOfSat = 0
		for (lowestsat,) in cursor:
			print(lowestsat)
			totalSaturation += lowestsat
			numberOfSat += 1
		average = float(totalSaturation) / numberOfSat
		#print(average)

	
			
	# If everything is taken care by the above two methods, we can just save the xlsx in this method
	def generateReport(self):
		self.wb.save(filename = self.imei_num+"_Surveillance_Report"+".xlsx")
